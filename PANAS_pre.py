"""
Created on Fri Apr  7 15:14:42 2023

@author: Bodoo
"""
import os
import openpyxl

wb1 = openpyxl.load_workbook("data_acq.xlsx")
sheet1 = wb1.active
print(wb1)

Header = ["#", "", "Upset",	"Hostile",	"Alert",	"Ashamed",	"Inspired",	"Nervous",	"Determined",   "Attentive",   "Afraid",   "Active"
]
sheet1.append(Header)


files = os.listdir("E:\Old Fear Extinction Task Data\csv_acq")

files_xlsx = [f for f in files if f[4] == 'c']


row1 = 3
col = 1

for f in files_xlsx:
    #data = pd.read_excel(f, usecols='AL', nrows=11)
    wb = openpyxl.load_workbook(f)
    sheet = wb.active
    #title= sheet1.cell(row1, col)    
    #title.value = f
    df = list()
    for i in range(2,14): 
        df.append(sheet.cell(row=i,column=38).value)
    sf = list()
    sf.append(f)
    ff = sf + df
    print(ff)
    #row1 += 1
    sheet1.append(ff)
wb1.save('data_acq.xlsx')

