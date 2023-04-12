# -*- coding: utf-8 -*-
"""
Created on Fri Apr  7 15:14:42 2023

@author: Bodoo
"""
import os
import openpyxl

wb1 = openpyxl.load_workbook("data_acq.xlsx")
sheet1 = wb1.active
print(wb1)

Header = ["#",	"state_anxiety",	"valence_yellow",	"valence_blue",	"arousal_yellow",	"arousal_blue", "contingency", "correct"]
sheet1.append(Header)


files = os.listdir("E:\Old Fear Extinction Task Data\Ext_xlsx")

files_xlsx = [f for f in files if f[4] == 'c']


row1 = 3
col = 1

for f in files_xlsx:
    #data = pd.read_excel(f, usecols='AL', nrows=11)
    wb = openpyxl.load_workbook(f)
    sheet = wb.active
    #title= sheet1.cell(row1, col)    
    #title.value = f
    df = list()
    
    df.append(f)
    
    df.append(sheet.cell(row=73,column=58).value)
    
    
    
    if sheet.cell(row=73,column=2).value == "yellow_sqr.png":
       df.append(sheet.cell(row=73,column=60).value) 
    if sheet.cell(row=74,column=2).value == "yellow_sqr.png":
       df.append(sheet.cell(row=74,column=60).value) 
    if sheet.cell(row=73,column=2).value == "blue_sqr.png":
       df.append(sheet.cell(row=73,column=60).value) 
    if sheet.cell(row=74,column=2).value == "blue_sqr.png":
       df.append(sheet.cell(row=74,column=60).value) 
       
       
       
    if sheet.cell(row=73,column=2).value == "yellow_sqr.png":
       df.append(sheet.cell(row=73,column=62).value) 
    if sheet.cell(row=74,column=2).value == "yellow_sqr.png":
       df.append(sheet.cell(row=74,column=62).value) 
    if sheet.cell(row=73,column=2).value == "blue_sqr.png":
       df.append(sheet.cell(row=73,column=62).value) 
    if sheet.cell(row=74,column=2).value == "blue_sqr.png":
       df.append(sheet.cell(row=74,column=62).value) 

    df.append(sheet.cell(row=75,column=64).value)
    
    df.append(sheet.cell(row=75,column=65).value)
    #for i in range(62,74): 
        #df.append(sheet.cell(row=i,column=56).value)

    print(df)
    #row1 += 1
    sheet1.append(df)
wb1.save('data.xlsx')

